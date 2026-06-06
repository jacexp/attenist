"""
SHEET-SCOPED MATCHING VERIFICATION
Tests that search, match, and correction are restricted to active sheet only.
"""
import os
import sys
import logging
from pathlib import Path
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(message)s")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.models import Employee
from services.ocr.validation_service import OCRValidationService
from database.database_service import DatabaseService

db = DatabaseService()
svc = OCRValidationService(db)

results = []
passed = 0
failed = 0

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
        results.append(f"  [PASS] {name}")
    else:
        failed += 1
        results.append(f"  [FAIL] {name}")
    if detail:
        results.append(f"         {detail}")

# Verify test workbook has employees across multiple sheets
all_emps = db.search_employees_as_objects("", 1000)
sheets = set(e.sheet_name for e in all_emps)

print(f"\nDatabase: {len(all_emps)} employees across {len(sheets)} sheets: {sheets}")

# Count employees per sheet
for s in sorted(sheets):
    count = sum(1 for e in all_emps if e.sheet_name == s)
    print(f"  {s}: {count} employees")

# Pick 3 sheets for cross-sheet testing
sheet_list = sorted(s for s in sheets if s and s != "None")
if len(sheet_list) >= 3:
    SHEET_A = sheet_list[0]
    SHEET_B = sheet_list[1]
    SHEET_C = sheet_list[2]
else:
    SHEET_A = sheet_list[0] if sheet_list else "Shif (2)"
    SHEET_B = sheet_list[1] if len(sheet_list) > 1 else SHEET_A
    SHEET_C = SHEET_A

# Pick one employee from each sheet
def get_emp_from_sheet(sheet_name):
    for e in all_emps:
        if e.sheet_name == sheet_name:
            return e
    return None

emp_a = get_emp_from_sheet(SHEET_A)
emp_b = get_emp_from_sheet(SHEET_B) if SHEET_B != SHEET_A else None
emp_c = get_emp_from_sheet(SHEET_C) if SHEET_C != SHEET_A and SHEET_C != SHEET_B else None

print(f"\nTest employees:")
print(f"  Sheet A '{SHEET_A}': {emp_a.employee_id if emp_a else 'N/A'}")
if emp_b:
    print(f"  Sheet B '{SHEET_B}': {emp_b.employee_id} {emp_b.name}")
if emp_c:
    print(f"  Sheet C '{SHEET_C}': {emp_c.employee_id} {emp_c.name}")


# ═══════════════════════════════════════════════════════════════════════
# TEST 1: search_employees_for_manual_match — sheet scoping
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST 1: search_employees_for_manual_match sheet scoping")
print("=" * 60)

if emp_a and emp_b:
    # Search for emp_b's ID while scoped to SHEET_A
    cross_sheet_id = emp_b.employee_id
    results_sheet_a = svc.search_employees_for_manual_match(
        cross_sheet_id, sheet_name=SHEET_A, limit=100
    )

    test(f"Search emp '{cross_sheet_id}' scoped to '{SHEET_A}' returns only '{SHEET_A}' employees",
         all(e.sheet_name == SHEET_A for e in results_sheet_a),
         f"returned {len(results_sheet_a)} results, "
         f"sheets={set(e.sheet_name for e in results_sheet_a)}")

    # Now search for the same ID while scoped to SHEET_B
    results_sheet_b = svc.search_employees_for_manual_match(
        cross_sheet_id, sheet_name=SHEET_B, limit=100
    )

    test(f"Search emp '{cross_sheet_id}' scoped to '{SHEET_B}' returns only '{SHEET_B}' employees",
         all(e.sheet_name == SHEET_B for e in results_sheet_b),
         f"returned {len(results_sheet_b)} results, "
         f"sheets={set(e.sheet_name for e in results_sheet_b)}")

    # emp_b should appear in SHEET_B results
    emp_b_found = any(e.employee_id == cross_sheet_id for e in results_sheet_b)
    test(f"emp '{cross_sheet_id}' found in '{SHEET_B}' results", emp_b_found)

    # emp_b should NOT appear in SHEET_A results (different employee ID on SHEET_A)
    emp_b_found_a = any(e.employee_id == cross_sheet_id for e in results_sheet_a)
    test(f"emp '{cross_sheet_id}' NOT found in '{SHEET_A}' results (different sheet)",
         not emp_b_found_a)

# Test with empty query (all employees)
print("\n--- Empty query (return all) ---")
results_all_a = svc.search_employees_for_manual_match("", sheet_name=SHEET_A, limit=1000)
test(f"Empty query scoped to '{SHEET_A}': all results are '{SHEET_A}'",
     all(e.sheet_name == SHEET_A for e in results_all_a),
     f"returned {len(results_all_a)} results")

# If multiple sheets exist, verify no cross-contamination
if len(sheet_list) >= 2:
    results_all_b = svc.search_employees_for_manual_match("", sheet_name=SHEET_B, limit=1000)
    a_ids = set(e.employee_id for e in results_all_a)
    b_ids = set(e.employee_id for e in results_all_b)
    overlap = a_ids & b_ids
    test(f"No employee ID overlap between '{SHEET_A}' and '{SHEET_B}'",
         not overlap,
         f"overlapping IDs: {overlap}" if overlap else "no overlap")


# ═══════════════════════════════════════════════════════════════════════
# TEST 2: _find_exact_match — sheet scoping
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST 2: _find_exact_match sheet scoping")
print("=" * 60)

if emp_a:
    # Find emp_a with its own sheet
    match = svc._find_exact_match(emp_a.employee_id, sheet_name=SHEET_A)
    test(f"_find_exact_match: emp '{emp_a.employee_id}' found in own sheet '{SHEET_A}'",
         match is not None and match.employee_id == emp_a.employee_id)

if emp_a and emp_b and SHEET_A != SHEET_B:
    # Try to find emp_b while scoped to SHEET_A — should fail
    cross_match = svc._find_exact_match(emp_b.employee_id, sheet_name=SHEET_A)
    test(f"_find_exact_match: emp '{emp_b.employee_id}' NOT found in wrong sheet '{SHEET_A}'",
         cross_match is None,
         f"got={cross_match.employee_id if cross_match else None}")


# ═══════════════════════════════════════════════════════════════════════
# TEST 3: manual_correction — sheet validation
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST 3: manual_correction sheet validation")
print("=" * 60)

from services.ocr.validation_service import OCRValidationResult, OCRStatus

if emp_b and emp_a and SHEET_A != SHEET_B:
    # Try to correct a result using emp_b (from SHEET_B) while active sheet is SHEET_A
    dummy_result = OCRValidationResult(
        ocr_id="TEST",
        ocr_name="TEST",
        status=OCRStatus.UNMATCHED,
    )

    try:
        svc.manual_correction(
            dummy_result,
            selected_employee=emp_b,
            sheet_name=SHEET_A,  # active sheet is A, but emp is from B
        )
        test(f"manual_correction: REJECTED cross-sheet correction (emp from {SHEET_B}, active {SHEET_A})",
             False, "Should have raised ValueError")
    except ValueError as e:
        test(f"manual_correction: REJECTED cross-sheet correction (emp from {SHEET_B}, active {SHEET_A})",
             True, str(e))
    except Exception as e:
        test(f"manual_correction: REJECTED cross-sheet correction",
             False, f"Unexpected exception: {e}")

    # Now try with same-sheet correction — should pass
    same_result = OCRValidationResult(
        ocr_id="TEST",
        ocr_name="TEST",
        status=OCRStatus.UNMATCHED,
    )
    try:
        svc.manual_correction(
            same_result,
            selected_employee=emp_a,
            sheet_name=SHEET_A,  # emp_a is from SHEET_A
        )
        test(f"manual_correction: ALLOWED same-sheet correction (emp from {SHEET_A}, active {SHEET_A})",
             same_result.matched_employee == emp_a,
             f"matched={same_result.matched_employee.employee_id if same_result.matched_employee else None}")
    except Exception as e:
        test(f"manual_correction: ALLOWED same-sheet correction",
             False, f"Unexpected exception: {e}")


# ═══════════════════════════════════════════════════════════════════════
# TEST 4: find_possible_matches — sheet scoping
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST 4: find_possible_matches sheet scoping")
print("=" * 60)

if emp_a:
    matches = svc.find_possible_matches(
        emp_a.employee_id, emp_a.name, SHEET_A, limit=10
    )
    test(f"find_possible_matches for '{SHEET_A}' returns only '{SHEET_A}' employees",
         all(m["employee"].sheet_name == SHEET_A for m in matches),
         f"returned {len(matches)} matches, "
         f"sheets={set(m['employee'].sheet_name for m in matches) if matches else 'none'}")

    if matches:
        test(f"find_possible_matches: emp '{emp_a.employee_id}' found in results",
             any(m["employee"].employee_id == emp_a.employee_id for m in matches))


# ═══════════════════════════════════════════════════════════════════════
# TEST 5: AttendanceService.mark() — sheet validation
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("TEST 5: AttendanceService.mark() sheet validation")
print("=" * 60)

from services.attendance_service import AttendanceService
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = SHEET_A
ws.cell(row=5, column=1, value=1)
ws.cell(row=emp_a.row, column=2, value=emp_a.employee_id)
ws.cell(row=emp_a.row, column=3, value=emp_a.name)
ws.cell(row=emp_a.row, column=4, value=emp_a.rank)
ws.cell(row=emp_a.row, column=6, value="WO")  # source cell

att_svc = AttendanceService(wb, [emp_a], {1: 1})

# Same-sheet mark — should pass
try:
    att_svc.mark(emp_a, 1, "A", SHEET_A)
    test(f"mark(): SAME-SHEET write allowed (emp from '{SHEET_A}', active '{SHEET_A}')", True)
except Exception as e:
    test(f"mark(): SAME-SHEET write allowed", False, f"Unexpected exception: {e}")

# Cross-sheet mark — should fail
if emp_b:
    try:
        att_svc.mark(emp_b, 1, "A", SHEET_A)
        test(f"mark(): REJECTED cross-sheet write (emp from '{SHEET_B}', active '{SHEET_A}')",
             False, "Should have raised ValueError")
    except ValueError as e:
        test(f"mark(): REJECTED cross-sheet write (emp from '{SHEET_B}', active '{SHEET_A}')",
             "Sheet mismatch" in str(e), str(e))
    except Exception as e:
        test(f"mark(): REJECTED cross-sheet write",
             False, f"Unexpected exception: {e}")


# ═══════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("SHEET-SCOPED MATCHING VERIFICATION SUMMARY")
print("=" * 60)
for r in results:
    print(r)

total = passed + failed
print(f"\n{'='*60}")
print(f"RESULTS: {passed}/{total} passed, {failed}/{total} failed")
print(f"{'='*60}")

sys.exit(0 if failed == 0 else 1)
