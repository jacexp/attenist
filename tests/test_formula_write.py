"""
Unit tests for formula-aware attendance write fix.

Cases:
  1. Direct formula ref: M13 = =F13 → write to F13, preserve =F13
  2. Chain formula ref:  T13 = =M13, M13 = =F13 → write to F13, preserve both
  3. Complex formula:    =COUNTIF(...) → abort with error
  4. Circular reference: A1 = =B1, B1 = =A1 → abort with error
"""
from pathlib import Path
from openpyxl import Workbook
import pytest

from core.models import Employee
from services.attendance_service import AttendanceService


# ─── Helpers ──────────────────────────────────────────────────────

def make_workbook():
    """Create a minimal workbook mirroring the real Shif (2) structure."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shif (2)"

    # Date header row
    for day, col in [(1,5),(2,6),(3,7),(4,8),(5,9),(6,10),(7,11),
                     (8,12),(9,13),(10,14),(11,15),(12,16)]:
        ws.cell(row=5, column=col, value=day)

    # Employee row
    ws.cell(row=13, column=1, value=7)   # serial
    ws.cell(row=13, column=2, value="BK447")
    ws.cell(row=13, column=3, value="BULBUL KUMARI")
    ws.cell(row=13, column=4, value="LGA")

    return wb


def make_employee():
    return Employee(
        employee_id="BK447",
        name="BULBUL KUMARI",
        rank="LGA",
        sheet_name="Shif (2)",
        row=13,
    )


def make_service(wb):
    employees = [make_employee()]
    dates = {11: 13, 18: 20}  # Day 11 -> col M(13), Day 18 -> col T(20)
    return AttendanceService(wb, employees, dates)


# ─── Case 1: Simple Direct Formula ───────────────────────────────

class TestDirectFormula:
    """M13 = =F13  →  write to F13, preserve =F13 in M13."""

    def test_resolves_to_source_cell(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")   # F13 = source data
        ws.cell(row=13, column=13).value = "=F13"  # M13 = formula
        svc = make_service(wb)

        svc.mark(make_employee(), 11, "AB", "Shif (2)")

        # Check M13 formula preserved
        m13 = ws.cell(row=13, column=13)
        assert m13.value == "=F13", f"M13 formula destroyed: {m13.value!r}"
        assert m13.data_type == 'f', f"M13 data_type changed: {m13.data_type!r}"

        # Check F13 source updated
        f13 = ws.cell(row=13, column=6)
        assert f13.value == "AB", f"F13 not updated: {f13.value!r}"

    def test_old_value_from_source(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")   # F13
        ws.cell(row=13, column=13).value = "=F13"  # M13
        svc = make_service(wb)

        old = svc.mark(make_employee(), 11, "AB", "Shif (2)")

        # old_value should be the source cell's old value
        assert old == "WO", f"old_value not from source: {old!r}"

    def test_non_formula_cell_unaffected(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")   # F13
        ws.cell(row=13, column=13).value = "=F13"  # M13
        svc = make_service(wb)

        svc.mark(make_employee(), 11, "AB", "Shif (2)")
        _, svc2 = make_service(wb), None  # fresh service

        # Save and re-open
        tmp = Path("/tmp/test_direct.xlsx")
        svc.save(tmp)
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp, data_only=False)
        ws2 = wb2["Shif (2)"]

        assert ws2.cell(row=13, column=13).value == "=F13", "M13 formula not preserved on disk"
        assert ws2.cell(row=13, column=6).value == "AB", "F13 value not persisted"
        tmp.unlink(missing_ok=True)


# ─── Case 2: Chain Formula ───────────────────────────────────────

class TestChainFormula:
    """T13 = =M13, M13 = =F13  →  write to F13, preserve both formulas."""

    def test_resolves_full_chain(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")    # F13 = source
        ws.cell(row=13, column=13).value = "=F13"  # M13 = formula -> F13
        ws.cell(row=13, column=20).value = "=M13"  # T13 = formula -> M13
        svc = make_service(wb)
        # Add Day 18 -> col 20 for the service
        svc.dates[18] = 20

        svc.mark(make_employee(), 18, "AB", "Shif (2)")

        # Check T13 formula preserved
        t13 = ws.cell(row=13, column=20)
        assert t13.value == "=M13", f"T13 formula destroyed: {t13.value!r}"
        assert t13.data_type == 'f', f"T13 data_type changed"

        # Check M13 formula preserved
        m13 = ws.cell(row=13, column=13)
        assert m13.value == "=F13", f"M13 formula destroyed: {m13.value!r}"
        assert m13.data_type == 'f', f"M13 data_type changed"

        # Check F13 source updated
        f13 = ws.cell(row=13, column=6)
        assert f13.value == "AB", f"F13 not updated: {f13.value!r}"

    def test_old_value_from_final_source(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")     # F13
        ws.cell(row=13, column=13).value = "=F13"  # M13
        ws.cell(row=13, column=20).value = "=M13"  # T13
        svc = make_service(wb)
        svc.dates[18] = 20

        old = svc.mark(make_employee(), 18, "AB", "Shif (2)")

        # old_value should be the final source cell's old value
        assert old == "WO", f"old_value not from final source: {old!r}"


# ─── Case 3: Complex Formula ─────────────────────────────────────

class TestComplexFormula:
    """=COUNTIF(...)  →  abort with clear error."""

    def test_raises_error(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")
        ws.cell(row=13, column=13).value = "=COUNTIF(F13:F13,\"A\")"
        svc = make_service(wb)

        with pytest.raises(ValueError) as exc:
            svc.mark(make_employee(), 11, "AB", "Shif (2)")

        assert "Complex formula" in str(exc.value)
        assert "COUNTIF" in str(exc.value)

    def test_raises_on_if(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=13).value = '=IF(E15="","",E15)'
        svc = make_service(wb)

        with pytest.raises(ValueError) as exc:
            svc.mark(make_employee(), 11, "AB", "Shif (2)")
        assert "Complex formula" in str(exc.value)

    def test_raises_on_sum(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=13).value = "=SUM(F13:G13)"
        svc = make_service(wb)

        with pytest.raises(ValueError) as exc:
            svc.mark(make_employee(), 11, "AB", "Shif (2)")
        assert "Complex formula" in str(exc.value)


# ─── Case 4: Circular Reference ──────────────────────────────────

class TestCircularReference:
    """A1 = =B1, B1 = =A1  →  abort with clear error."""

    def test_direct_circular(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6).value = "=F13"  # F13 -> F13 (self-loop)
        ws.cell(row=13, column=13).value = "=F13"  # M13 -> F13 (which is circular)
        svc = make_service(wb)

        with pytest.raises(ValueError) as exc:
            svc.mark(make_employee(), 11, "AB", "Shif (2)")
        assert "Circular reference" in str(exc.value)

    def test_indirect_circular(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=20).value = "=M13"  # T13 -> M13
        ws.cell(row=13, column=13).value = "=T13"  # M13 -> T13 (cycle)
        ws.cell(row=13, column=6, value="WO")       # F13 = source
        svc = make_service(wb)
        svc.dates[18] = 20

        with pytest.raises(ValueError) as exc:
            svc.mark(make_employee(), 11, "AB", "Shif (2)")
        assert "Circular reference" in str(exc.value)


# ─── Case 5: Invalid Reference ───────────────────────────────────

class TestInvalidReference:
    """Reference to non-existent cell/row → abort."""

    def test_invalid_row_zero(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=13).value = "=A0"
        svc = make_service(wb)

        with pytest.raises(ValueError) as exc:
            svc.mark(make_employee(), 11, "AB", "Shif (2)")
        assert "reference" in str(exc.value).lower() or "invalid" in str(exc.value).lower()


# ─── Integration: Save and Reopen ────────────────────────────────

class TestSaveAndReopen:
    """After write + save, formula is preserved on disk."""

    def test_save_preserves_formulas(self):
        wb = make_workbook()
        ws = wb["Shif (2)"]
        ws.cell(row=13, column=6, value="WO")
        ws.cell(row=13, column=13).value = "=F13"   # M13
        ws.cell(row=13, column=20).value = "=M13"   # T13
        svc = make_service(wb)
        svc.dates[18] = 20

        svc.mark(make_employee(), 11, "AB", "Shif (2)")  # writes to F13

        tmp = Path("/tmp/test_save_formula.xlsx")
        svc.save(tmp)

        from openpyxl import load_workbook
        wb2 = load_workbook(tmp, data_only=False)
        ws2 = wb2["Shif (2)"]

        assert ws2.cell(row=13, column=6).value == "AB", "F13 wrong on disk"
        assert ws2.cell(row=13, column=13).value == "=F13", "M13 formula not preserved"
        assert ws2.cell(row=13, column=13).data_type == 'f', "M13 not formula on disk"
        assert ws2.cell(row=13, column=20).value == "=M13", "T13 formula not preserved"

        tmp.unlink(missing_ok=True)
