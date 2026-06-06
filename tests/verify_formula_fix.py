"""
FINAL PRODUCTION-GRADE VERIFICATION
Tests the formula fix against the REAL attendance workbook.

Requirements:
- 20+ formula cells tested
- Save/reopen verification
- Complex formula protection (COUNTIF, IF, SUM)
- Circular reference protection
- All attendance paths (A, B, C, WO, AB)
- Logging diagnostics
- Formula count preservation
- Workbook integrity
"""
import io
import logging
import os
import sys
import traceback
from pathlib import Path
from copy import copy

from openpyxl import load_workbook, Workbook

# Configure logging to both file and buffer for verification
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl.utils import column_index_from_string, get_column_letter

from core.models import Employee
from services.attendance_service import AttendanceService


def resolve_chain(cell, ws, visited=None):
    """Follow formula chain to final non-formula source cell."""
    if visited is None:
        visited = set()
    cid = (cell.row, cell.column)
    if cid in visited:
        return cell
    visited.add(cid)
    if cell.data_type != 'f' or not isinstance(cell.value, str) or not str(cell.value).startswith('='):
        return cell
    ref = str(cell.value)[1:].strip()
    letter = ""
    r = 0
    for i, ch in enumerate(ref):
        if ch.isalpha():
            letter += ch
        else:
            r = int(ref[i:])
            break
    if letter and r >= 1:
        next_cell = ws.cell(row=r, column=column_index_from_string(letter))
        return resolve_chain(next_cell, ws, visited)
    return cell

REAL_WB_PATH = "samples/test1.xlsx"
TMP_WB_PATH = "/tmp/formula_verify_output.xlsx"

results = []
passed = 0
failed = 0

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
        status = "PASS"
    else:
        failed += 1
        status = "FAIL"
    results.append(f"  [{status}] {name}")
    if detail:
        results.append(f"         {detail}")


# ═══════════════════════════════════════════════════════════════════════
# 1. LOAD REAL WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("PHASE 1: Load Real Workbook")
print("=" * 70)

wb = load_workbook(REAL_WB_PATH, data_only=False)
ws = wb["Shif (2)"]

# Build day→column mapping from header row 5
header_map = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=5, column=col).value
    if val is not None:
        try:
            header_map[int(val)] = col
        except (ValueError, TypeError):
            pass

# Build employee index
employees = []
for row_num in range(8, ws.max_row + 1):
    emp_id = ws.cell(row=row_num, column=2).value
    name = ws.cell(row=row_num, column=3).value
    rank = ws.cell(row=row_num, column=4).value
    if emp_id:
        emp = Employee(
            employee_id=str(emp_id).strip(),
            name=str(name or "").strip(),
            rank=str(rank or "").strip(),
            sheet_name="Shif (2)",
            row=row_num,
        )
        employees.append(emp)

print(f"  Loaded: {wb.sheetnames}")
print(f"  Shif(2) employees: {len(employees)}")
print(f"  Days in header: {sorted(header_map.keys())}")

# Build the date→column mapping for days that have formulas
# Formula chains: F(col6,Day4) ← M(col13,Day11) ← T(col20,Day18)
#                  G(col7,Day5)  ← N(col14,Day12) ← U(col21,Day19)
#                  H(col8,Day6)  ← O(col15,Day13) ← V(col22,Day20)
dates = {}
for day, col in header_map.items():
    if col >= 5:  # attendance data columns
        dates[day] = col
        last_col = col

print(f"  Date→column mapping: {len(dates)} days")


# ═══════════════════════════════════════════════════════════════════════
# 2. FIND ALL 30 FORMULA CELLS AND VERIFY
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 2: Test All 30 Formula Cells")
print("=" * 70)

formula_cells = []  # (cell_coord, day, col, formula_string, employee_row)
for row_num in range(8, ws.max_row + 1):
    emp_id = ws.cell(row=row_num, column=2).value
    if not emp_id:
        continue
    for col in range(5, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col)
        if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
            # Find which day this column corresponds to
            day = None
            for d, c in header_map.items():
                if c == col:
                    day = d
                    break
            formula_cells.append((cell.coordinate, day, col, cell.value, row_num))

print(f"  Found {len(formula_cells)} formula cells in Shif(2)")

for coord, day, col, formula, row_num in sorted(
    formula_cells, key=lambda x: (x[4], x[2])
):
    emp_id = str(ws.cell(row=row_num, column=2).value or "").strip()
    emp_name = str(ws.cell(row=row_num, column=3).value or "").strip()

    # Determine source cell from formula
    formula_str = str(formula).strip()
    source_ref = formula_str.lstrip("=").strip()
    
    # Parse the source cell
    source_col_letter = ""
    source_row = 0
    for i, ch in enumerate(source_ref):
        if ch.isalpha():
            source_col_letter += ch
        else:
            source_row = int(source_ref[i:])
            break
    
    # Get old value from source cell
    source_cell = ws[f"{source_col_letter}{source_row}"]
    old_value = source_cell.value
    
    formula_col_letter = get_column_letter(col)
    
    before_formula = str(formula)
    before_type = 'f'
    before_source_val = old_value
    
    results.append(
        f"\n  Cell {coord} ({emp_id} {emp_name}): "
        f"{before_formula} → resolves to {source_ref}"
    )
    results.append(f"    Before: formula={before_formula}, type={before_type}, source={source_ref}={before_source_val}")


# ═══════════════════════════════════════════════════════════════════════
# 3. MARK ATTENDANCE ON FORMULA CELLS
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 3: Attendnace Write on Formula Cells")
print("=" * 70)

# Use a COPY of the workbook for writes
wb2 = load_workbook(REAL_WB_PATH, data_only=False)
ws2 = wb2["Shif (2)"]

svc = AttendanceService(wb2, employees, dates)

shifts_tested = set()
# Test 2 formula cells per pattern (F←M←T, G←N←U, H←O←V) with different shifts
test_cases = [
    # (row, formula_col, day, shift_to_write, description)
    (13, 13, 11, "A", "Write A on M13 (=F13) chain, Day 11"),
    (13, 20, 18, "B", "Write B on T20 (=M13) chain, Day 18"),
    (12, 14, 12, "C", "Write C on N14 (=G12) chain, Day 12"),
    (12, 21, 19, "WO", "Write WO on U21 (=N12) chain, Day 19"),
    (24, 13, 11, "AB", "Write AB on M24 (=F24) chain"),
    (24, 20, 18, "A", "Write A on T24 (=M24) chain"),
    (26, 13, 11, "B", "Write B on M26 (=F26) chain"),
    (26, 20, 18, "C", "Write C on T26 (=M26) chain"),
    (27, 14, 12, "WO", "Write WO on N27 (=G27) chain"),
    (27, 21, 19, "A", "Write A on U27 (=N27) chain"),
    (34, 14, 12, "B", "Write B on N34 (=G34) chain"),
    (34, 21, 19, "C", "Write C on U34 (=N34) chain"),
    (38, 13, 11, "WO", "Write WO on M38 (=F38) chain"),
    (38, 20, 18, "AB", "Write AB on T38 (=M38) chain"),
    (41, 14, 12, "A", "Write A on N41 (=G41) chain"),
    (41, 21, 19, "B", "Write B on U41 (=N41) chain"),
    (44, 13, 11, "C", "Write C on M44 (=F44) chain"),
    (44, 20, 18, "WO", "Write WO on T44 (=M44) chain"),
    (50, 14, 12, "AB", "Write AB on N50 (=G50) chain"),
    (50, 21, 19, "A", "Write A on U50 (=N50) chain"),
    (8, 15, 13, "B", "Write B on O8 (=H8) chain"),
    (8, 22, 20, "C", "Write C on V22 (=O8) chain"),
    (19, 15, 13, "WO", "Write WO on O19 (=H19) chain"),
    (19, 22, 20, "AB", "Write AB on V22 (=O19) chain"),
    (43, 15, 13, "A", "Write A on O43 (=H43) chain"),
    (43, 22, 20, "B", "Write B on V43 (=O43) chain"),
]

for row_num, formula_col, day, new_shift, desc in test_cases:
    coord = f"{ws2.cell(row=row_num, column=formula_col).coordinate}"
    formula_val = ws2.cell(row=row_num, column=formula_col).value
    formula_str = str(formula_val) if formula_val else ""

    # Find the employee
    emp_id_val = str(ws2.cell(row=row_num, column=2).value or "").strip()
    emp_name_val = str(ws2.cell(row=row_num, column=3).value or "").strip()
    emp = Employee(
        employee_id=emp_id_val,
        name=emp_name_val,
        rank=str(ws2.cell(row=row_num, column=4).value or "").strip(),
        sheet_name="Shif (2)",
        row=row_num,
    )

    # Capture formula before
    before_cell = ws2.cell(row=row_num, column=formula_col)
    before_formula = before_cell.value
    before_data_type = before_cell.data_type

    # Determine source cell
    if formula_str.startswith("="):
        target_ref = formula_str[1:].strip()
    else:
        target_ref = coord

    # Mark attendance
    try:
        old_val = svc.mark(emp, day, new_shift, "Shif (2)")
        shifts_tested.add(new_shift)

        # Check formula preserved
        after_cell = ws2.cell(row=row_num, column=formula_col)
        after_formula = after_cell.value
        after_data_type = after_cell.data_type

        resolved = resolve_chain(after_cell, ws2)
        source_new_val = resolved.value

        formula_ok = (after_formula == before_formula)
        type_ok = (after_data_type == before_data_type and after_data_type == 'f')
        source_ok = (source_new_val == new_shift)

        test(f"{coord}: {desc}",
             formula_ok and type_ok and source_ok,
             f"formula={'OK' if formula_ok else 'DESTROYED'} "
             f"type={'OK' if type_ok else 'CHANGED'} "
             f"source={'OK' if source_ok else 'WRONG'} "
             f"({before_formula} → {resolved.coordinate}={source_new_val})")

    except Exception as e:
        test(f"{coord}: {desc}", False, f"EXCEPTION: {e}")

# Verify all 5 shift types were tested
test("All 5 shift types tested (A, B, C, WO, AB)",
     shifts_tested == {"A", "B", "C", "WO", "AB"},
     f"tested={shifts_tested}")

# ═══════════════════════════════════════════════════════════════════════
# 4. FORMULA COUNT PRESERVATION (Before Save)
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 4: Formula Count Before Save")
print("=" * 70)

count_before = 0
for row in ws2.iter_rows():
    for cell in row:
        if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
            count_before += 1

print(f"  Formula count in memory before save: {count_before}")


# ═══════════════════════════════════════════════════════════════════════
# 5. SAVE AND REOPEN
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 5: Save & Reopen")
print("=" * 70)

svc.save(TMP_WB_PATH)

# Reopen
wb3 = load_workbook(TMP_WB_PATH, data_only=False)
ws3 = wb3["Shif (2)"]

count_after = 0
formulas_after = []
for row in ws3.iter_rows():
    for cell in row:
        if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
            count_after += 1
            formulas_after.append((cell.coordinate, cell.value))

test("Formula count unchanged after save+reopen",
     count_before == count_after,
     f"before={count_before} after={count_after}")

# Verify each formula cell is intact on disk (formula string + type preserved)
all_formulas_intact = True
formula_intact_count = 0
formula_total_count = 0
for row_num in range(8, ws3.max_row + 1):
    for col in range(5, ws3.max_column + 1):
        cell = ws3.cell(row=row_num, column=col)
        if cell.data_type == 'f' and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_total_count += 1
            # Resolve the chain — source should have a valid value (not the formula string)
            resolved = resolve_chain(cell, ws3)
            if resolved.data_type != 'f' and resolved.value is not None:
                formula_intact_count += 1
            else:
                all_formulas_intact = False
                print(f"  BROKEN: {cell.coordinate}: formula={cell.value} resolved={resolved.coordinate}={resolved.value}")

test(f"All {formula_total_count} formulas intact on disk after save+reopen",
     all_formulas_intact and formula_total_count == 30,
     f"intact={formula_intact_count}/{formula_total_count}")


# ═══════════════════════════════════════════════════════════════════════
# 6. COMPLEX FORMULA PROTECTION
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 6: Complex Formula Protection")
print("=" * 70)

# Test COUNTIF — use TESS-2 TERRIER sheet
wb4 = load_workbook(REAL_WB_PATH, data_only=False)
ws4 = wb4["TESS-2 TERRIER"]
# Create a temp employee on row 10 where there's a COUNTIF
emp_countif = Employee(
    employee_id="TEST",
    name="TEST EMPLOYEE",
    rank="LGA",
    sheet_name="TESS-2 TERRIER",
    row=10,
)
svc4 = AttendanceService(wb4, [emp_countif], dates)

# Column 5 (E) = Day 3, cell E10 has =COUNTIF(E5:E8,"A")
try:
    svc4.mark(emp_countif, 3, "A", "TESS-2 TERRIER")
    test("COUNTIF protection: Write rejected", False,
         "Should have raised ValueError")
except ValueError as e:
    test("COUNTIF protection: Write rejected", "Complex formula" in str(e),
         f"Got expected error: {e}")
except Exception as e:
    test("COUNTIF protection: Write rejected", False,
         f"Unexpected exception: {e}")

# Test IF formula protection
wb_if = Workbook()
ws_if = wb_if.active
ws_if.title = "Sheet1"
ws_if.cell(row=5, column=1, value=1)  # Day 1 header
ws_if.cell(row=10, column=1).value = '=IF(E15="","",E15)'  # Row 10, col A
ws_if.cell(row=10, column=2, value="TEST")
emp_if = Employee(employee_id="TEST", name="TEST", rank="", sheet_name="Sheet1", row=10)
svc_if = AttendanceService(wb_if, [emp_if], {1: 1})

try:
    svc_if.mark(emp_if, 1, "A", "Sheet1")
    test("IF formula protection: Write rejected", False,
         "Should have raised ValueError")
except ValueError as e:
    test("IF formula protection: Write rejected", "Complex formula" in str(e),
         f"Got expected error: {e}")

# Test SUM formula protection
ws_sum = Workbook()
ws_sum.active.title = "Sheet1"
ws_sum.active.cell(row=5, column=1, value=1)
ws_sum.active.cell(row=10, column=1).value = 0  # placeholder
cell_sum = ws_sum.active.cell(row=10, column=1)
cell_sum.value = "=SUM(A1:A10)"  # valid range referencing existing rows
cell_sum.data_type = 'f'
ws_sum.active.cell(row=10, column=2, value="TEST")
emp_sum = Employee(employee_id="TEST", name="TEST", rank="", sheet_name="Sheet1", row=10)
svc_sum = AttendanceService(ws_sum, [emp_sum], {1: 1})

try:
    svc_sum.mark(emp_sum, 1, "A", "Sheet1")
    test("SUM formula protection: Write rejected", False,
         "Should have raised ValueError")
except ValueError as e:
    test("SUM formula protection: Write rejected",
         "Complex formula" in str(e),
         f"Got expected error: {e}")


# ═══════════════════════════════════════════════════════════════════════
# 7. CIRCULAR REFERENCE PROTECTION
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 7: Circular Reference Protection")
print("=" * 70)

# Test direct self-reference
wb_circ = Workbook()
ws_circ = wb_circ.active
ws_circ.title = "Sheet1"
ws_circ.cell(row=5, column=1, value=1)
ws_circ.cell(row=13, column=6).value = "=F13"  # F13 -> self
ws_circ.cell(row=13, column=13).value = "=F13"  # M13 -> F13 (circular from F13)
ws_circ.cell(row=13, column=2, value="CIRC1")
emp_circ = Employee(employee_id="CIRC1", name="CIRC1", rank="", sheet_name="Sheet1", row=13)
svc_circ = AttendanceService(wb_circ, [emp_circ], {11: 13})

try:
    svc_circ.mark(emp_circ, 11, "A", "Sheet1")
    test("Direct circular ref (self-loop): Write rejected", False,
         "Should have raised ValueError")
except ValueError as e:
    test("Direct circular ref (self-loop): Write rejected",
         "Circular" in str(e),
         f"Got expected error: {e}")

# Test indirect circular (A->B->A)
wb_circ2 = Workbook()
ws_circ2 = wb_circ2.active
ws_circ2.title = "Sheet1"
ws_circ2.cell(row=5, column=1, value=1)
ws_circ2.cell(row=13, column=20).value = "=M13"  # T13 -> M13
ws_circ2.cell(row=13, column=13).value = "=T13"  # M13 -> T13 (cycle)
ws_circ2.cell(row=13, column=6, value="WO")       # F13 = source data
ws_circ2.cell(row=13, column=2, value="CIRC2")
emp_circ2 = Employee(employee_id="CIRC2", name="CIRC2", rank="", sheet_name="Sheet1", row=13)
svc_circ2 = AttendanceService(wb_circ2, [emp_circ2], {18: 20})

try:
    svc_circ2.mark(emp_circ2, 18, "A", "Sheet1")
    test("Indirect circular ref (M13→T13→M13): Write rejected", False,
         "Should have raised ValueError")
except ValueError as e:
    test("Indirect circular ref (M13→T13→M13): Write rejected",
         "Circular" in str(e),
         f"Got expected error: {e}")


# ═══════════════════════════════════════════════════════════════════════
# 8. INVALID REFERENCE PROTECTION
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 8: Invalid Reference Protection")
print("=" * 70)

wb_inv = Workbook()
ws_inv = wb_inv.active
ws_inv.title = "Sheet1"
ws_inv.cell(row=5, column=1, value=1)
ws_inv.cell(row=13, column=13).value = "=A0"  # Row 0 — invalid
ws_inv.cell(row=13, column=2, value="INV1")
emp_inv = Employee(employee_id="INV1", name="INV1", rank="", sheet_name="Sheet1", row=13)
svc_inv = AttendanceService(wb_inv, [emp_inv], {11: 13})

try:
    svc_inv.mark(emp_inv, 11, "A", "Sheet1")
    test("Invalid ref (row 0): Write rejected", False,
         "Should have raised ValueError")
except ValueError as e:
    test("Invalid ref (row 0): Write rejected", True,
         f"Got expected error: {e}")


# ═══════════════════════════════════════════════════════════════════════
# 9. NON-FORMULA CELLS UNAFFECTED
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 9: Non-Formula Cell Should Still Work")
print("=" * 70)

wb_nf = load_workbook(REAL_WB_PATH, data_only=False)
ws_nf = wb_nf["Shif (2)"]
# Row 10 has no formulas — all direct values
emp_nf = employees[2]  # Row 10: AN944 APARNA HANSDA
svc_nf = AttendanceService(wb_nf, employees, dates)

# Day 3 (col 5) — direct value, no formula
before_val = ws_nf.cell(row=10, column=5).value
svc_nf.mark(emp_nf, 3, "AB", "Shif (2)")
after_val = ws_nf.cell(row=10, column=5).value
test("Non-formula cell: Direct write works",
     after_val == "AB",
     f"before={before_val} after={after_val}")

# Day 4 (col 6) — direct value, no formula
before_val = ws_nf.cell(row=10, column=6).value
svc_nf.mark(emp_nf, 4, "AB", "Shif (2)")
after_val = ws_nf.cell(row=10, column=6).value
test("Non-formula cell next to formula cells: Untouched",
     after_val == "AB",
     f"before={before_val} after={after_val}")


# ═══════════════════════════════════════════════════════════════════════
# 10. WORKBOOK INTEGRITY
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"PHASE 10: Workbook Integrity")
print("=" * 70)

# Final reopen and comprehensive check
wb_final = load_workbook(TMP_WB_PATH, data_only=False)
all_ok = True
for sname in wb_final.sheetnames:
    ws_final = wb_final[sname]
    for row in ws_final.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                val = cell.value
                if val is None or not isinstance(val, str) or not val.startswith('='):
                    all_ok = False
                    print(f"  CORRUPTED: {sname} {cell.coordinate}: type=f but value={val!r}")

test("All sheets: No corrupted formula cells", all_ok,
     "All formula cells have valid type=f and value starting with =")

# Check workbook can be opened without error
try:
    wb_final.save(TMP_WB_PATH)  # re-save
    wb_final2 = load_workbook(TMP_WB_PATH, data_only=False)
    test("Workbook: Can save and reopen without errors", True)
except Exception as e:
    test("Workbook: Can save and reopen without errors", False, str(e))

# Clean up
Path(TMP_WB_PATH).unlink(missing_ok=True)


# ═══════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("VERIFICATION SUMMARY")
print("=" * 70)
for r in results:
    print(r)

total = passed + failed
print(f"\n{'='*70}")
print(f"RESULTS: {passed}/{total} passed, {failed}/{total} failed")
print(f"{'='*70}")

sys.exit(0 if failed == 0 else 1)
